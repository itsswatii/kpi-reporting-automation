import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from google import genai
import os
from datetime import datetime

# load the data
df = pd.read_csv('data/superstore_sales.csv', encoding='latin1')
print(df.shape)
print(df.head())

# fix dates
df['Order Date'] = pd.to_datetime(df['Order Date'], dayfirst=True)
df['Ship Date'] = pd.to_datetime(df['Ship Date'], dayfirst=True)
df['Year'] = df['Order Date'].dt.year
df['Month'] = df['Order Date'].dt.month
df['YearMonth'] = df['Order Date'].dt.to_period('M')

# quick check for nulls
print(df.isnull().sum())

# basic KPIs
total_sales = df['Sales'].sum()
total_orders = df['Order ID'].nunique()
avg_order_value = total_sales / total_orders
total_customers = df['Customer ID'].nunique()

print(f"\nTotal Sales: ${total_sales:,.2f}")
print(f"Total Orders: {total_orders:,}")
print(f"Avg Order Value: ${avg_order_value:,.2f}")
print(f"Total Customers: {total_customers:,}")

# breakdowns
sales_by_category = df.groupby('Category')['Sales'].sum().sort_values(ascending=False)
sales_by_region = df.groupby('Region')['Sales'].sum().sort_values(ascending=False)
sales_by_segment = df.groupby('Segment')['Sales'].sum().sort_values(ascending=False)
monthly_sales = df.groupby('YearMonth')['Sales'].sum().reset_index()
monthly_sales['YearMonth'] = monthly_sales['YearMonth'].astype(str)
top_products = df.groupby('Product Name')['Sales'].sum().sort_values(ascending=False).head(10)

print("\nSales by Category:")
print(sales_by_category)
print("\nSales by Region:")
print(sales_by_region)

# visuals
plt.figure(figsize=(8, 5))
sns.barplot(x=sales_by_category.index, y=sales_by_category.values, palette='Set2')
plt.title('Total Sales by Category')
plt.xlabel('Category')
plt.ylabel('Total Sales ($)')
plt.tight_layout()
plt.savefig('visuals/sales_by_category.png')
plt.show()

plt.figure(figsize=(8, 5))
sns.barplot(x=sales_by_region.index, y=sales_by_region.values, palette='coolwarm')
plt.title('Total Sales by Region')
plt.xlabel('Region')
plt.ylabel('Total Sales ($)')
plt.tight_layout()
plt.savefig('visuals/sales_by_region.png')
plt.show()

plt.figure(figsize=(14, 5))
plt.plot(monthly_sales['YearMonth'], monthly_sales['Sales'], marker='o', color='steelblue')
plt.title('Monthly Sales Trend')
plt.xlabel('Month')
plt.ylabel('Sales ($)')
plt.xticks(rotation=90)
plt.tight_layout()
plt.savefig('visuals/monthly_sales_trend.png')
plt.show()

plt.figure(figsize=(8, 5))
sns.barplot(x=sales_by_segment.index, y=sales_by_segment.values, palette='Blues_d')
plt.title('Total Sales by Customer Segment')
plt.xlabel('Segment')
plt.ylabel('Total Sales ($)')
plt.tight_layout()
plt.savefig('visuals/sales_by_segment.png')
plt.show()

# call claude api to generate an executive summary
kpi_summary = f"""
Total Sales: ${total_sales:,.2f}
Total Orders: {total_orders:,}
Average Order Value: ${avg_order_value:,.2f}
Total Customers: {total_customers:,}

Sales by Category:
{sales_by_category.to_string()}

Sales by Region:
{sales_by_region.to_string()}

Sales by Segment:
{sales_by_segment.to_string()}

Top 3 Products:
{top_products.head(3).to_string()}
"""

top_category = sales_by_category.index[0]
top_region = sales_by_region.index[0]
top_segment = sales_by_segment.index[0]

prompt = f"""You are a senior business analyst. Based on the KPI data below, write a concise executive summary (3-4 paragraphs) covering:
1. Overall business performance
2. Top performing categories and regions
3. Key trends and anomalies
4. 2-3 actionable recommendations

KPI Data:
{kpi_summary}

Keep it professional and data-driven."""

try:
    client = genai.Client(api_key=os.environ["GEMINI_API_KEY"])
    response = client.models.generate_content(
        model="gemini-2.0-flash",
        contents=prompt
    )
    ai_narrative = response.text
    print("\nAI Narrative (Gemini):")
except Exception as e:
    print(f"\nGemini unavailable ({e.__class__.__name__}), using rule-based narrative.")
    ai_narrative = f"""
The business recorded total sales of ${total_sales:,.2f} across {total_orders:,} orders from {total_customers:,} unique customers, with an average order value of ${avg_order_value:,.2f}.

{top_category} emerged as the top-performing category. Regionally, {top_region} led all regions in total sales, indicating strong market presence in that area.

The {top_segment} segment drove the highest sales volume among customer segments, representing the core customer base.

Recommendations: (1) Increase investment in {top_category} and {top_region}. (2) Develop targeted promotions for lower-performing regions. (3) Monitor monthly trends to capitalize on seasonal peaks.
"""

print(ai_narrative)


# generate excel report
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "Executive Summary"

ws1['A1'] = "AI-Powered KPI Report - Superstore Sales"
ws1['A1'].font = Font(bold=True, size=16, color="FFFFFF")
ws1['A1'].fill = PatternFill("solid", fgColor="2F5496")
ws1['A1'].alignment = Alignment(horizontal='center')
ws1.merge_cells('A1:D1')

ws1['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
ws1['A2'].font = Font(italic=True, size=10)
ws1.merge_cells('A2:D2')

ws1['A4'] = "KEY PERFORMANCE INDICATORS"
ws1['A4'].font = Font(bold=True, size=12, color="2F5496")

for col, header in enumerate(['Metric', 'Value'], 1):
    cell = ws1.cell(row=5, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal='center')

kpis = [
    ('Total Sales', f"${total_sales:,.2f}"),
    ('Total Orders', f"{total_orders:,}"),
    ('Avg Order Value', f"${avg_order_value:,.2f}"),
    ('Total Customers', f"{total_customers:,}"),
]

for row, (metric, value) in enumerate(kpis, 6):
    ws1.cell(row=row, column=1, value=metric).font = Font(bold=True)
    ws1.cell(row=row, column=2, value=value)

ws1['A11'] = "AI-GENERATED EXECUTIVE SUMMARY"
ws1['A11'].font = Font(bold=True, size=12, color="2F5496")
ws1['A12'] = ai_narrative
ws1['A12'].alignment = Alignment(wrap_text=True)
ws1.merge_cells('A12:D25')
ws1.row_dimensions[12].height = 300

ws1.column_dimensions['A'].width = 30
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 20
ws1.column_dimensions['D'].width = 20

# sales by category sheet
ws2 = wb.create_sheet("Sales by Category")
ws2.append(['Category', 'Total Sales ($)'])
for cat, sales in sales_by_category.items():
    ws2.append([cat, round(sales, 2)])

# sales by region sheet
ws3 = wb.create_sheet("Sales by Region")
ws3.append(['Region', 'Total Sales ($)'])
for region, sales in sales_by_region.items():
    ws3.append([region, round(sales, 2)])

# monthly trend sheet
ws4 = wb.create_sheet("Monthly Trend")
ws4.append(['Month', 'Total Sales ($)'])
for _, r in monthly_sales.iterrows():
    ws4.append([r['YearMonth'], round(r['Sales'], 2)])

wb.save('output/KPI_Report_Superstore.xlsx')
print("\nExcel report saved to output/KPI_Report_Superstore.xlsx")
